#!/usr/bin/env python3
"""Build an AI-facing experiment workbook for the B02 paper."""
from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path("/private/tmp/b02_ai_table_pull")
OUT_XLSX = Path("/private/tmp/b02_ai_table_pull/B02_AI_Experiment_Master_Table_20260715.xlsx")
OUT_MD = Path("/private/tmp/b02_ai_table_pull/B02_AI_Experiment_Master_Table_README.md")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="D9EAD3")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def safe_sheet_name(name: str, used: set[str]) -> str:
    name = re.sub(r"[^A-Za-z0-9_]+", "_", name).strip("_")[:31] or "sheet"
    base = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def autosize(ws, max_w=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_w)


def write_rows(wb, title: str, rows: list[dict], description: str, used: set[str]):
    ws = wb.create_sheet(safe_sheet_name(title, used))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = description
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    if not rows:
        ws["A4"] = "(no rows)"
        return ws
    keys = []
    seen = set()
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                keys.append(k)
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=4, column=c, value=k)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r, row in enumerate(rows, 5):
        for c, k in enumerate(keys, 1):
            v = row.get(k, "")
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A5"
    autosize(ws)
    return ws


def read_csv(path: Path) -> list[dict]:
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def flatten_json_record(x, prefix="") -> dict:
    if isinstance(x, dict):
        out = {}
        for k, v in x.items():
            key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
            if isinstance(v, (dict, list)):
                out[key] = json.dumps(v, ensure_ascii=False)
            else:
                out[key] = v
        return out
    return {"value": x}


def read_json_rows(path: Path) -> list[dict]:
    with open(path) as f:
        data = json.load(f)
    if isinstance(data, list):
        return [flatten_json_record(x) for x in data]
    if isinstance(data, dict):
        # Prefer key-value rows for metadata-like dicts.
        return [{"key": k, "value": json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v} for k, v in data.items()]
    return [{"value": data}]


def num(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default


def load_optional_csv(rel: str) -> list[dict]:
    p = ROOT / rel
    return read_csv(p) if p.exists() else []


def load_optional_json(rel: str) -> list[dict]:
    p = ROOT / rel
    return read_json_rows(p) if p.exists() else []


def build_scorecard() -> list[dict]:
    final_claims = load_optional_csv("full_aggregates/final_claim_support.csv")
    supp_claims = load_optional_csv("supplement_aggregates/Table_G_Updated_Claims.csv")
    live = load_optional_csv("supplemental_20260715_results/live_vllm_affinity_workload.csv")
    adm = load_optional_csv("supplemental_20260715_results/admission_ablation.csv")
    stale = load_optional_csv("supplemental_20260715_results/staleness_validation.csv")
    event = load_optional_csv("supplemental_20260715_results/rate_controlled_event_driven.csv")

    high_live = [r for r in live if r.get("locality") == "high"]
    high_coarse = next((r for r in high_live if r.get("policy") == "coarse"), {})
    high_sk2 = next((r for r in high_live if r.get("policy") == "sketch" and r.get("K") == "2"), {})
    high_exact = next((r for r in high_live if r.get("policy") == "exact"), {})
    low_live = [r for r in live if r.get("locality") == "low"]
    low_coarse = next((r for r in low_live if r.get("policy") == "coarse"), {})
    low_sk8 = next((r for r in low_live if r.get("policy") == "sketch" and r.get("K") == "8"), {})

    high_adm = [r for r in adm if r.get("locality") == "high" and r.get("policy") == "demand_aware"]
    med_adm = [r for r in adm if r.get("locality") == "medium" and r.get("policy") == "demand_aware"]
    stale_max_unsafe = max([num(r.get("unsafe_reuse_rate")) for r in stale], default=0)
    low_churn = [r for r in event if r.get("churn_events_per_inst_s") == "0.1" and num(r.get("budget_bytes_per_inst_s")) >= 256]
    median_red10 = sorted([num(r.get("reduction_vs_periodic_10Hz_x")) for r in low_churn])
    median_red50 = sorted([num(r.get("reduction_vs_periodic_50Hz_x")) for r in low_churn])
    med10 = median_red10[len(median_red10)//2] if median_red10 else 0
    med50 = median_red50[len(median_red50)//2] if median_red50 else 0

    return [
        {
            "claim_id": "C1",
            "paper_claim": "Exact/Rich reusable-state visibility is much more expensive than coarse load state.",
            "support_level": "Strong",
            "best_evidence": "Prior vLLM experiments show Rich/Coarse state-size ratios of about 8.9x, 11.1x, and 22.3x.",
            "primary_sheets": "StateSize_*, full_state_view_size, supplement_Table_A/D",
            "writing_guidance": "Use as the main motivation for a bounded state interface. Avoid implying Rich always improves quality.",
        },
        {
            "claim_id": "C2",
            "paper_claim": "Coarse/load-only dispatch misses prefix affinity under agentic locality.",
            "support_level": "Strong for locality workloads; weak for low-locality workloads",
            "best_evidence": f"Live T4 high-locality: coarse candidate_hit={high_coarse.get('candidate_hit_rate')} reuse={high_coarse.get('observed_reuse_hit_rate')} vs sketch K=2 candidate_hit={high_sk2.get('candidate_hit_rate')} reuse={high_sk2.get('observed_reuse_hit_rate')}. Prior clean experiments also show Sketch cache hit above Coarse.",
            "primary_sheets": "live_vllm_affinity_workload, supplement_Table_A, full_dispatch_quality",
            "writing_guidance": "Frame as workload-dependent: prefix locality is required. Low-locality rows show limited benefit.",
        },
        {
            "claim_id": "C3",
            "paper_claim": "A bounded Sketch can recover much of exact affinity value with bounded metadata.",
            "support_level": "Moderate to strong",
            "best_evidence": f"Live T4 high-locality: sketch K=2 TTFT p95={high_sk2.get('ttft_p95_ms')} ms, exact p95={high_exact.get('ttft_p95_ms')} ms, coarse p95={high_coarse.get('ttft_p95_ms')} ms. Admission simulation: demand-aware high-locality saved-vs-exact ranges {', '.join(str(r.get('saved_vs_exact_ratio')) for r in high_adm[:4])}.",
            "primary_sheets": "live_vllm_affinity_workload, admission_ablation, ksweep_all_summaries",
            "writing_guidance": "Say bounded top-K is effective; do not say the current demand-aware heuristic is universally optimal because LRU/coverage sometimes win.",
        },
        {
            "claim_id": "C4",
            "paper_claim": "Rate-controlled event-driven updates avoid N*K*f periodic metadata scaling.",
            "support_level": "Strong model/simulation support",
            "best_evidence": f"Supplemental token-bucket simulation at low churn has median reductions of about {med10:.1f}x vs 10Hz and {med50:.1f}x vs 50Hz periodic reporting.",
            "primary_sheets": "rate_controlled_event_driven, eventdriven_all_results",
            "writing_guidance": "Describe as control-plane simulation/model evidence unless a live transport implementation is added.",
        },
        {
            "claim_id": "C5",
            "paper_claim": "Stale metadata cannot cause unsafe KV reuse because the owner validates before reuse.",
            "support_level": "Strong protocol simulation support",
            "best_evidence": f"Staleness injection up to 20% stale entries produced max unsafe_reuse_rate={stale_max_unsafe}.",
            "primary_sheets": "staleness_validation",
            "writing_guidance": "Use for correctness/safety of the interface, not for performance speedup.",
        },
        {
            "claim_id": "C6",
            "paper_claim": "Benefits are workload-dependent.",
            "support_level": "Strong",
            "best_evidence": f"Live low-locality: coarse reuse={low_coarse.get('observed_reuse_hit_rate')} vs sketch K=8 reuse={low_sk8.get('observed_reuse_hit_rate')}; benefits are much smaller than high-locality.",
            "primary_sheets": "live_vllm_affinity_workload, admission_ablation",
            "writing_guidance": "Include this explicitly as a limitation and scope condition.",
        },
    ]


def main():
    wb = Workbook()
    wb.remove(wb.active)
    used = set()

    readme_rows = [
        {"item": "Purpose", "value": "AI-facing master table for writing the experiment section of Cost-Aware State Interfaces for LLM Request Dispatch."},
        {"item": "Scope", "value": "Combines prior T4/vLLM experiments, N=256 mock dispatcher data, K-sweep, event-driven simulations, and 2026-07-15 supplemental experiments."},
        {"item": "Strong claims", "value": "Rich state is costly; coarse state misses affinity under locality; bounded sketch recovers useful affinity; event-driven metadata reduces traffic; validation prevents unsafe stale reuse."},
        {"item": "Weak/conditional claims", "value": "TTFT gains are workload/model dependent; current K-sweep same_inst_step_ratio is not discriminative; synthetic simulations support interface behavior but are not full serving benchmarks."},
        {"item": "How to use", "value": "Start with Evidence_Scorecard and Claim_Map. Then pull exact numbers from raw data sheets. Use Data_Dictionary for field meanings and Caveats for limitations."},
    ]
    write_rows(wb, "AI_Readme", readme_rows, "Read this first.", used)
    write_rows(wb, "Evidence_Scorecard", build_scorecard(), "Claim-by-claim evidence assessment and writing guidance.", used)

    claim_map = [
        {"paper_section": "Introduction/Motivation", "claim": "State interface matters, not just dispatch policy.", "use_sheets": "Evidence_Scorecard, full_state_view_size, supplement_Table_A, live_vllm_affinity_workload", "suggested_text": "A load-only interface hides reusable prefix KV; exact affinity exposes it but inflates metadata."},
        {"paper_section": "Background 2.2", "claim": "Workflow ID is not itself reusable state.", "use_sheets": "staleness_validation, admission_ablation", "suggested_text": "Advertised entries identify physically resident, compatible KV resources; stale or missing entries fall back."},
        {"paper_section": "Design 3.3", "claim": "Demand-aware top-K admission is a bounded heuristic.", "use_sheets": "admission_ablation", "suggested_text": "Demand-aware top-K preserves substantial exact-affinity value at K=4-16, but heuristic choice is workload-sensitive."},
        {"paper_section": "Design 3.4", "claim": "Event-driven dissemination is rate controlled.", "use_sheets": "rate_controlled_event_driven, eventdriven_all_results", "suggested_text": "Under low churn, event-driven updates reduce bytes/s by orders of magnitude relative to periodic N*K*f reporting."},
        {"paper_section": "Evaluation", "claim": "Live T4 sanity check.", "use_sheets": "live_vllm_affinity_workload", "suggested_text": "On running Qwen2.5-1.5B vLLM endpoints, high-locality Sketch/Exact routing substantially increases reuse and lowers TTFT relative to coarse round-robin."},
        {"paper_section": "Limitations", "claim": "Benefits require prefix locality and larger prefill costs.", "use_sheets": "live_vllm_affinity_workload, admission_ablation, Prior_AllTTFT", "suggested_text": "Low-locality workloads and small-model TTFT measurements show smaller or noisy gains; the main contribution is the bounded interface."},
    ]
    write_rows(wb, "Claim_Map", claim_map, "Where each dataset should be used in the paper.", used)

    dictionary = [
        {"field": "policy", "meaning": "Dispatch/state-interface variant: coarse/load-only, rich/exact, sketch/top-K, random/lru/coverage ablations.", "notes": "Compare policies only within the same experiment and workload."},
        {"field": "K", "meaning": "Maximum advertised affinity entries per instance or per simulated instance.", "notes": "full means exact/unbounded for that experiment."},
        {"field": "locality", "meaning": "Prefix popularity distribution: high/medium/low reuse.", "notes": "Used in supplemental experiments to expose workload sensitivity."},
        {"field": "hit_rate/cache_hit/candidate_hit", "meaning": "How often the dispatcher or serving path found reusable prefix-affinity candidates.", "notes": "Definitions differ by experiment; use sheet descriptions."},
        {"field": "observed_reuse_hit_rate", "meaning": "In live vLLM workload, fraction of requests whose prefix had already been sent to the selected instance.", "notes": "Dispatcher-level proxy for prefix-cache opportunity."},
        {"field": "ttft_p50_ms/ttft_p95_ms", "meaning": "Time to first token in milliseconds.", "notes": "Small T4/Qwen2.5-1.5B experiments can be noisy; use mainly as supporting evidence."},
        {"field": "metadata_snapshot_bytes", "meaning": "Estimated bytes for one state snapshot/index advertisement.", "notes": "Control-plane size model, not payload tokens."},
        {"field": "reduction_vs_periodic_*_x", "meaning": "Bytes/s ratio of periodic full reporting over event-driven updates.", "notes": "Higher is better for event-driven."},
        {"field": "unsafe_reuse_rate", "meaning": "Fraction of requests that incorrectly reused incompatible/missing KV after validation.", "notes": "Should be zero; validates stale-metadata safety claim."},
        {"field": "saved_vs_exact_ratio", "meaning": "Simulated saved-prefill latency captured by bounded policy divided by exact affinity.", "notes": "A bounded-interface quality metric."},
    ]
    write_rows(wb, "Data_Dictionary", dictionary, "Field meanings for AI paper writing.", used)

    caveats = [
        {"caveat": "Current K-sweep same_inst_step_ratio is not discriminative", "detail": "Existing ksweep results show same_inst_step_ratio=1.0 even for coarse, so use TTFT/metadata and supplemental admission data instead of claiming K quality from that metric."},
        {"caveat": "Demand-aware is not always best", "detail": "Supplemental admission ablation shows LRU or coverage can win for some generated distributions. Write demand-aware as a concrete heuristic, not a proven optimizer."},
        {"caveat": "Event-driven is simulation/model evidence", "detail": "It supports scaling and budget arguments but is not a full live network transport implementation."},
        {"caveat": "N=256 mock is dispatcher/control-plane only", "detail": "Use it for collection and decision overhead, not as real 256-GPU serving."},
        {"caveat": "TTFT on small T4 model is noisy", "detail": "Use live T4 results as sanity checks. Stronger TTFT claims would need larger model/longer context, e.g. A800."},
    ]
    write_rows(wb, "Caveats", caveats, "Limitations and safe wording.", used)

    datasets = [
        ("exp_part_a_real_serving", "experiments_aggregates/part_a_real_serving_results.csv", "Early real-serving T4 experiment results."),
        ("exp_part_b_stress", "experiments_aggregates/part_b_stress_test_results.csv", "Early stress test results."),
        ("exp_state_freq", "experiments_aggregates/state_frequency_summary.csv", "State frequency / maintenance summary."),
        ("exp_state_size", "experiments_aggregates/state_size_summary.csv", "State size summary."),
        ("full_dispatch_quality", "full_aggregates/dispatch_quality.csv", "Full experiment dispatch quality."),
        ("full_final_claims", "full_aggregates/final_claim_support.csv", "Prior final claim support table."),
        ("full_maintenance_cost", "full_aggregates/maintenance_cost.csv", "Full experiment maintenance cost."),
        ("full_state_view_size", "full_aggregates/state_view_size.csv", "Full experiment state view sizes."),
        ("full_stress_test", "full_aggregates/stress_test.csv", "Full experiment stress results."),
        ("full_tradeoff", "full_aggregates/tradeoff_summary.csv", "Full experiment quality/cost tradeoff."),
        ("supp_Table_A_clean", "supplement_aggregates/Table_A_Clean_Tradeoff.csv", "Supplement clean tradeoff table."),
        ("supp_Table_B_context", "supplement_aggregates/Table_B_Long_Context.csv", "Long context sensitivity."),
        ("supp_Table_C_workflow", "supplement_aggregates/Table_C_Workflow_Length.csv", "Workflow length sensitivity."),
        ("supp_Table_D_rich_size", "supplement_aggregates/Table_D_Rich_Size_Breakdown.csv", "Rich state size breakdown."),
        ("supp_Table_E_ablation", "supplement_aggregates/Table_E_Sketch_Ablation.csv", "Sketch ablation."),
        ("supp_Table_F_stress", "supplement_aggregates/Table_F_Stress_Target_vs_Achieved.csv", "Stress target vs achieved."),
        ("supp_Table_G_claims", "supplement_aggregates/Table_G_Updated_Claims.csv", "Supplement updated claims."),
        ("n256_baseline", "n256_aggregates/Table_Baseline_ByPolicy.csv", "N=256 mock dispatcher baseline."),
        ("n256_fail", "n256_aggregates/Table_FailInjection.csv", "N=256 mock failure injection."),
        ("n256_final_claims", "n256_aggregates/Table_FinalClaims_N256.csv", "N=256 final claims."),
        ("n256_policy_agg", "n256_aggregates/Table_PerPolicy_Aggregate.csv", "N=256 per-policy aggregate."),
        ("supp_admission", "supplemental_20260715_results/admission_ablation.csv", "2026-07-15 CPU-only admission heuristic ablation."),
        ("supp_staleness", "supplemental_20260715_results/staleness_validation.csv", "2026-07-15 stale metadata validation."),
        ("supp_event_budget", "supplemental_20260715_results/rate_controlled_event_driven.csv", "2026-07-15 token-bucket event-driven dissemination simulation."),
        ("supp_key_findings", "supplemental_20260715_results/supplemental_key_findings.csv", "2026-07-15 compact key findings."),
        ("live_vllm_affinity", "supplemental_20260715_results/live_vllm_affinity_workload.csv", "2026-07-15 live T4/vLLM prefix-affinity workload."),
    ]
    for name, rel, desc in datasets:
        p = ROOT / rel
        if p.exists():
            write_rows(wb, name, read_csv(p), desc + f" Source: {rel}", used)

    json_datasets = [
        ("ksweep_all_summaries", "ksweep_results/all_summaries.json", "Existing K-sweep summaries. Use with caveat about same_inst_step_ratio."),
        ("eventdriven_all_results", "eventdriven/all_results.json", "Existing event-driven vs periodic simulation."),
        ("supp_run_metadata", "supplemental_20260715_results/run_metadata.json", "Supplemental run metadata."),
    ]
    for name, rel, desc in json_datasets:
        p = ROOT / rel
        if p.exists():
            write_rows(wb, name, read_json_rows(p), desc + f" Source: {rel}", used)

    manifest_rows = []
    for p in sorted(ROOT.rglob("*")):
        if p.is_file():
            manifest_rows.append({
                "relative_path": str(p.relative_to(ROOT)),
                "bytes": p.stat().st_size,
                "include_in_workbook": "yes" if p.suffix.lower() in [".csv", ".json"] else "metadata/report only",
            })
    write_rows(wb, "Source_Manifest", manifest_rows, "All files pulled into the AI table build directory.", used)

    wb.save(OUT_XLSX)

    with open(OUT_MD, "w") as f:
        f.write("# B02 AI Experiment Master Table\n\n")
        f.write(f"Workbook: `{OUT_XLSX.name}`\n\n")
        f.write("Use `Evidence_Scorecard` first. It tells a paper-writing AI which claims are strong, which are conditional, and which sheets contain the supporting numbers.\n\n")
        f.write("Recommended experiment-section framing:\n\n")
        f.write("1. Metadata cost: Rich/exact state is 8.9x-22.3x larger than coarse.\n")
        f.write("2. Dispatch quality: coarse misses prefix affinity under locality; Sketch recovers much of the benefit.\n")
        f.write("3. Bounded K: top-K bounds metadata; K should be presented as an interface budget knob.\n")
        f.write("4. Freshness: event-driven updates and owner validation bound traffic and preserve correctness under stale metadata.\n")
        f.write("5. Scope: benefits depend on prefix locality and prefill cost; current T4 small-model TTFT is supporting, not definitive.\n")

    print(OUT_XLSX)
    print(OUT_MD)


if __name__ == "__main__":
    main()
